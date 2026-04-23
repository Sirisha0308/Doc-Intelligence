import streamlit as st
import os
from langchain_community.document_loaders import PyMuPDFLoader
from langchain_community.document_loaders import Docx2txtLoader
from langchain_community.document_loaders import TextLoader
from langchain_community.document_loaders import CSVLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import Chroma
from langchain_groq import ChatGroq
from dotenv import load_dotenv
load_dotenv()
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser
from langchain_core.messages import HumanMessage, AIMessage

# ── Page config ──────────────────────────────────────
st.set_page_config(
    page_title="Document Intelligence System",
    page_icon="docs",
    layout="centered"
)

st.title("Document Intelligence System")
st.caption("Upload any document and ask questions about it")

# ── Load models once (cached) ────────────────────────
@st.cache_resource
def load_embeddings():
    return HuggingFaceEmbeddings(
        model_name="sentence-transformers/all-MiniLM-L6-v2"
    )

@st.cache_resource
def load_llm():
    return ChatGroq(
        model="llama-3.1-8b-instant",
        temperature=0
    )

# ── Check if file is an image ────────────────────────
def is_image_file(filename):
    image_extensions = [
        "jpg", "jpeg", "png", "gif",
        "bmp", "webp", "tiff", "svg"
    ]
    ext = filename.split(".")[-1].lower()
    return ext in image_extensions

# ── Get correct loader for file type ────────────────
def get_loader(file_path, file_type):
    if file_type == "pdf":
        return PyMuPDFLoader(file_path)
    elif file_type == "docx":
        return Docx2txtLoader(file_path)
    elif file_type == "txt":
        return TextLoader(file_path, encoding="utf-8")
    elif file_type == "csv":
        return CSVLoader(file_path)
    elif file_type in ["xlsx", "xls"]:
        return None
    else:
        return None
    
def load_excel(file_path):
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text += f"Sheet: {sheet}\n"
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(
                str(cell) for cell in row
                if cell is not None
            )
            if row_text.strip():
                text += row_text + "\n"
    from langchain_core.documents import Document
    return [Document(page_content=text)]

# ── Build vector store ───────────────────────────────
def build_vectorstore(file_path, file_type, embeddings):
    # handle Excel separately
    if file_type in ["xlsx", "xls"]:
        docs = load_excel(file_path)
    else:
        loader = get_loader(file_path, file_type)
        if loader is None:
            return None, 0, 0
        docs = loader.load()

    # filter empty pages
    docs = [doc for doc in docs
            if doc.page_content.strip() != ""]

    if len(docs) == 0:
        return None, 0, 0

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=500,
        chunk_overlap=50,
        separators=["\n\n", "\n", ". ", " ", ""]
    )
    chunks = splitter.split_documents(docs)

    if len(chunks) == 0:
        return None, 0, 0

    vectorstore = Chroma.from_documents(
    documents=chunks,
    embedding=embeddings
)
    return vectorstore, len(docs), len(chunks)

# ── Build RAG chain with conversation memory ─────────
def build_rag_chain(vectorstore, llm):
    retriever = vectorstore.as_retriever(
        search_kwargs={"k": 3}
    )

    # prompt now includes chat history
    prompt = PromptTemplate(
        template="""You are a helpful assistant.
Use only the context below to answer the question.
If the answer is not in the context, say
"I don't know based on the document."

Context from document:
{context}

Conversation so far:
{chat_history}

Question: {question}

Answer:""",
        input_variables=[
            "context", "chat_history", "question"
        ]
    )

    def format_docs(docs):
        return "\n\n".join(
            doc.page_content for doc in docs
        )

    # formats chat history into readable string
    def format_history(messages):
        if not messages:
            return "No previous conversation."
        history = ""
        for msg in messages:
            if isinstance(msg, HumanMessage):
                history += f"User: {msg.content}\n"
            elif isinstance(msg, AIMessage):
                history += f"Assistant: {msg.content}\n"
        return history

    def rag_with_memory(inputs):
        question = inputs["question"]
        chat_history = inputs["chat_history"]

        # get relevant docs
        relevant_docs = retriever.invoke(question)
        context = format_docs(relevant_docs)

        # format history
        history_text = format_history(chat_history)

        # run through prompt and llm
        chain = prompt | llm | StrOutputParser()
        answer = chain.invoke({
            "context": context,
            "chat_history": history_text,
            "question": question
        })
        return answer

    return rag_with_memory

# ── Session state setup ──────────────────────────────
if "messages" not in st.session_state:
    st.session_state.messages = []
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "rag_chain" not in st.session_state:
    st.session_state.rag_chain = None
if "current_file" not in st.session_state:
    st.session_state.current_file = None

# ── Load models ──────────────────────────────────────
embeddings = load_embeddings()
llm = load_llm()

# ── Sidebar ──────────────────────────────────────────
with st.sidebar:
    st.header("Upload Document")
    st.caption("Supported: PDF, DOCX, TXT, CSV")

    uploaded_file = st.file_uploader(
    "Choose a file",
    type=[
        "pdf", "docx", "txt", "csv", "xlsx", "xls",
        "jpg", "jpeg", "png", "gif", "bmp", "webp"
    ]
)

    if uploaded_file is not None:
        # check if user uploaded an image
        if is_image_file(uploaded_file.name):
            st.error(
                "This is an image file, not a document. "
                "Please upload a document such as "
                "a PDF, Word file, or text file."
            )

        elif uploaded_file.name != \
                st.session_state.current_file:
            file_type = \
                uploaded_file.name.split(".")[-1].lower()

            with st.spinner("Processing document..."):
                temp_path = \
                    f"./temp_{uploaded_file.name}"
                with open(temp_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                try:
                    vectorstore, num_pages, num_chunks\
                        = build_vectorstore(
                            temp_path,
                            file_type,
                            embeddings
                        )

                    if vectorstore is None:
                        st.error(
                            "Could not read this file. "
                            "It may be a scanned image "
                            "or empty document. "
                            "Please upload a different "
                            "document."
                        )
                    else:
                        st.session_state.vectorstore \
                            = vectorstore

                        rag_chain = build_rag_chain(
                            vectorstore, llm
                        )
                        st.session_state.rag_chain \
                            = rag_chain
                        st.session_state.current_file \
                            = uploaded_file.name
                        st.session_state.messages = []
                        st.session_state.chat_history \
                            = []

                        st.success("Document ready!")

                except Exception as e:
                    st.error(f"Error: {str(e)}")

                finally:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)

    if st.session_state.current_file:
        st.divider()
        st.write("Current document:")
        st.write(
            f"**{st.session_state.current_file}**"
        )

    st.divider()
    st.caption(
        "Built with LangChain + Groq + ChromaDB"
    )

# ── Main chat area ───────────────────────────────────
if st.session_state.rag_chain is None:
    st.info(
        "Upload a document from the sidebar "
        "to get started"
    )
else:
    # show chat history
    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            st.write(msg["content"])

    # chat input
    if question := st.chat_input(
        "Ask a question about your document..."
    ):
        # show user message
        with st.chat_message("user"):
            st.write(question)
        st.session_state.messages.append({
            "role": "user",
            "content": question
        })

        # get answer
        with st.chat_message("assistant"):
            with st.spinner("Thinking..."):
                try:
                    answer = st.session_state\
                        .rag_chain({
                            "question": question,
                            "chat_history":
                                st.session_state
                                .chat_history
                        })

                    st.write(answer)

                    # save to chat history for memory
                    st.session_state.chat_history\
                        .append(
                            HumanMessage(
                                content=question
                            )
                        )
                    st.session_state.chat_history\
                        .append(
                            AIMessage(content=answer)
                        )

                    st.session_state.messages.append({
                        "role": "assistant",
                        "content": answer
                    })

                except Exception as e:
                    st.error(
                        f"Error getting answer: "
                        f"{str(e)}"
                    )